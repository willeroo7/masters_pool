import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os

class ExcelHandler:
    def __init__(self):
        self.data_dir = 'data'
        os.makedirs(self.data_dir, exist_ok=True)

    def generate_report(self, scores_data):
        """
        Generate an Excel report from the scores data.
        Returns the path to the generated file.
        """
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Masters Tournament Scores"

        # Define headers
        headers = ['Position', 'Player Name', 'Total Score', 'Round 1', 'Round 2', 'Round 3', 'Round 4']
        
        # Style for headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Write data
        for row, player in enumerate(scores_data, 2):
            ws.cell(row=row, column=1, value=player['position'])
            ws.cell(row=row, column=2, value=player['name'])
            ws.cell(row=row, column=3, value=player['total_score'])
            ws.cell(row=row, column=4, value=player['rounds']['round1'])
            ws.cell(row=row, column=5, value=player['rounds']['round2'])
            ws.cell(row=row, column=6, value=player['rounds']['round3'])
            ws.cell(row=row, column=7, value=player['rounds']['round4'])

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"masters_scores_{timestamp}.xlsx"
        filepath = os.path.join(self.data_dir, filename)

        # Save the workbook
        wb.save(filepath)
        return filepath 